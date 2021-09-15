# coding=utf-8
import openpyxl

wb = openpyxl.load_workbook('test.xlsx')

# 获取表格的页数,通过index或者sheet名称获取
ws = wb.worksheets[1]  # index
ws = wb.get_sheet_by_name('test')  # sheet
data = ws.cell(1, 2).value
print(data)

# 获取所有的数据
rows = ws.rows

# 获取指定区域内的数据 大于第一行,第一列的数据
rows = ws.iter_rows(min_row=1, min_clo=2)
print(rows)

for row in rows:
    for c in row:
        print(c.value)
