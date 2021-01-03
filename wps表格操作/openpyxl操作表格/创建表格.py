# coding=utf-8
import openpyxl
from openpyxl.utils import get_column_letter

# 创建表格
wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0, title="test")
# 一行一行的插入数据,插到表格最下边
ws.append([1, 2, 3, 4, 5, 6])
a = [1, 2, 3]
ws.append(a)
# 给指定的位置插入数据
ws.cell(row=5, column=5).value = '12344'
# 指定位置插入空行,在第二行插入3行数据
ws.insert_rows(2, 3)
# 保存数据
wb.save('test.xlsx')
print(2222222222222)

