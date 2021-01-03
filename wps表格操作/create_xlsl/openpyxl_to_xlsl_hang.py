#coding=utf-8
from openpyxl import Workbook
import numpy as np

wb = Workbook()
ws = wb.create_sheet(index=0,title="test")
label = ['hahah','2','dawdfa',1,3]
features = [
           [0.1, 0.2, 0.3, 0.4, 0.5],
           [0.11, 0.21, 0.31, 0.41, 0.51],
           [0.6, 0.7, 0.8, 0.9, 1.00],
           [1.1, 1.2, 1.3, 1.4, 1.5],
           ]
ws.append(label)
for feature in features:
    ws.append(feature)
ws = wb.create_sheet(index=1,title="test1")
ws.append(label)
ws = wb.get_sheet_by_name('test')
ws.append(label)
wb.save("tests2.xlsx")
