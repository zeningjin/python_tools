#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2020/10/10 上午11:36
# @Author  : jinzening
# @File    : ZenTao_base.py
# @Software: PyCharm
import requests
import json
import datetime
from lxml import etree
from hashlib import md5
from openpyxl import Workbook
import logging

class Zentao(object):

    def __init__(self, host, account, password):
        self.host = host
        self.session_id = ''
        self.params = {}
        self.account = account
        self.password = password
        self._get_session_id()
        self.login()

    def _get_session_id(self):
        api_path = "/zentao/api-getsessionid.json"
        response = requests.get(self.host + api_path)
        result = self._get_zentao_result(response.json())
        self.session_id = result['sessionID']
        self.params[result['sessionName']] = self.session_id

    def login(self):
        api_path = '/zentao/user-login.json'
        data = {
            'account': self.account,
            'password': self.password,
        }
        result = requests.post(self.host + api_path, data=data, params=self.params).json()
        if result['status'] == 'success':
            print('zentao 登陆成功')
        else:
            print(result)

    @staticmethod
    def _get_zentao_result(result):
        if result['status'] == 'success' and md5(result['data'].encode()).hexdigest() == result['md5']:
            data = json.loads(result['data'])
            return data
        return result

    def get_project_task(self):
        path = '/zentao/project-all-all.html'
        res = self.zentao_get(path)
        ele = etree.HTML(res)
        project_ids = ele.xpath("//*[@id='projectTableList']/tr/td[1]/text()")
        return [project_id.strip() for project_id in project_ids]

    def get_task_list(self):
        project_ids = self.get_project_task()
        project_data = {}
        for project_id in project_ids:
            task_data = {}
            api_path = "/zentao/project-task-" + str(project_id) +"-all.html"
            res = self.zentao_get(api_path)
            ele = etree.HTML(res)
            # 获取项目的总数量
            all_count = ele.xpath("//*[@id='all']/span[2]/text()")[0]
            print(all_count)
            for j in range(1, int(int(all_count) / 100 + 0.99)):
                api_path = "/zentao/project-task-" + str(project_id) + "-all-0--" \
                           + str(all_count) + "-100-" + str(j) + ".html"
                print(self.host + api_path)
                res = self.zentao_get(api_path)
                ele = etree.HTML(res)
                for i in range(1, 100):
                    try:
                        task_id = ele.xpath("//*[@id='taskList']/tbody/tr["+str(i)+"]/td[1]/a/text()")[0]
                        task_name = ele.xpath("//*[@id='taskList']/tbody/tr["+str(i)+"]/td[2]/a/text()")[0]
                        complete_people = ele.xpath("//*[@id='taskList']/tbody/tr["+str(i)+"]/td[4]/a/span/text()")[0]
                        status = ele.xpath("//*[@id='taskList']/tbody/tr["+str(i)+"]/td[3]/span/text()")[0]
                        complete_time = ele.xpath("//*[@id='taskList']/tbody/tr["+str(i)+"]/td[5]/text()")[0]
                        product_line = ele.xpath("//*[@id='currentItem']/text()")[0]
                        is_subtask = ele.xpath("//*[@id='taskList']/tbody/tr["+str(i)+"]/td[2]/span/text()")
                        if is_subtask:
                            continue
                        print(task_id, task_name, complete_people, status, complete_time, product_line)
                        task_data[task_id] = {
                            'task_name': task_name,
                            'complete_people': complete_people,
                            'status': status,
                            'complete_time': complete_time,
                            'product_line': product_line,
                        }
                    except:
                        break
            project_data[project_id] = task_data
        print(project_data)
        return project_data

    def get_remark_by_task_id(self, task_id):
        """
        通过task_id获取该任务的备注
        :param task_id:
        :return:
        """
        path = '/zentao/task-view-' + str(task_id) + '.html'
        res = self.zentao_get(path)
        ele = etree.HTML(res)
        ########################## 之后需要修改哦
        pre_online_remaek = ele.xpath("//*[@id='actionbox']/div[2]/ol/li[2]/div/div/span/text()")
        online_remaek = ele.xpath("//*[@id='actionbox']/div[2]/ol/li[3]/div/div/span/text()")
        return pre_online_remaek, online_remaek

    def get_subtask_ids_by_task_id(self, task_id):
        """
        通过task_id获取该任务的任务详情
        :param task_id:
        :return:
        """
        path = '/zentao/task-view-' + str(task_id) + '.html'
        res = self.zentao_get(path)
        ele = etree.HTML(res)
        # 获取子任务的task_id
        develop_peoples = []
        project_list = []
        pre_online_remaek = {}
        online_remaek = {}
        online_data = {}
        subtask_ids = ele.xpath("//*[@id='mainContent']/div[1]/div[1]/div[2]/div[2]/table/tbody/tr/td[1]/text()")
        if subtask_ids:
            for subtask_id in subtask_ids:
                # 判断该子任务是否为测试类型任务及完成状态
                if self.get_pre_message_by_task_id(subtask_id):
                    pre_online_remaek[task_id] = {}
                    online_remaek[task_id] = {}
                    for subtask_id in subtask_ids:
                        # 判断上线情况
                        if self.get_message_by_task_id(subtask_id):
                            print(self.get_message_by_task_id(subtask_id))
                            online_data = self.get_message_by_task_id(subtask_id)
                        path = '/zentao/task-view-' + str(subtask_id) + '.html'
                        res = self.zentao_get(path)
                        ele = etree.HTML(res)
                        # 获取任务类型
                        task_type = ele.xpath("//*[@id='legendBasic']/table/tbody/tr[5]/td/text()")
                        if task_type == ['测试']:
                            test_people = ele.xpath("//*[@id='legendBasic']/table/tbody/tr[4]/td/text()")
                            if test_people:
                                pre_online_remaek[task_id]['test_people'] = test_people[0].split(' ')[0]
                            pre_online_remaek[task_id]['configure'] = '不涉及'
                            if ele.xpath("//*[@id='mainContent']/div[1]/div[1]/div/div[2]/text()"):
                                pre_online_remaek[task_id]['configure'] = \
                                    ele.xpath("//*[@id='mainContent']/div[1]/div[1]/div/div[2]/text()")[0].strip()
                        if task_type == ['开发']:
                            develop_people = ele.xpath("//*[@id='legendBasic']/table/tbody/tr[4]/td/text()")
                            if develop_people and not (develop_people[0] .split(' ')[0] in develop_peoples):
                                develop_peoples.append(develop_people[0].split(' ')[0])
                            project = ele.xpath("//*[@id='mainContent']/div[1]/div[1]/div/div[2]/span/text()")
                            if not project:
                                project = ele.xpath("//*[@id='mainContent']/div[1]/div[1]/div/div[2]/text()")
                            if project:
                                for i in project:
                                    i = i.strip()
                                    project_list.append(i)
        if develop_peoples and project_list:
            project_list = ",".join(set(project_list))
            develop_peoples = ",".join(set(develop_peoples))
            pre_online_remaek[task_id]['develop_people'] = develop_peoples
            pre_online_remaek[task_id]['project'] = project_list
        if online_data.get('status', '') == '通过':
            online_remaek = pre_online_remaek
            online_remaek[task_id]['result'] = '通过'
            pre_online_remaek = {}
        if online_data.get('status', '') == '未通过':
            online_remaek = pre_online_remaek
            online_remaek[task_id]['result'] = '未通过'
        return pre_online_remaek, online_remaek

    def get_pre_message_by_task_id(self, task_id):
        """
        判断测试子任务是否完成
        :param task_id:
        :return:
        """
        path = '/zentao/task-view-' + str(task_id) + '.html'
        res = self.zentao_get(path)
        ele = etree.HTML(res)
        # 获取任务类型
        task_type = ele.xpath("//*[@id='legendBasic']/table/tbody/tr[5]/td/text()")
        if task_type == ['测试']:
            task_status = ele.xpath("//*[@id='legendBasic']/table/tbody/tr[6]/td/span/text()")
            if task_status == [' 已完成']:
                return True
        return False

    def get_message_by_task_id(self, task_id):
        """
        判断上线子任务是否完成
        :param task_id:
        :return:
        """
        path = '/zentao/task-view-' + str(task_id) + '.html'
        res = self.zentao_get(path)
        ele = etree.HTML(res)
        data = {}
        try:
            # 获取任务类型
            task_type = ele.xpath("//*[@id='legendBasic']/table/tbody/tr[5]/td/text()")
            pro_time = ele.xpath("//*[@id='legendBasic']/table/tbody/tr[4]/td/text()")[0].split(' ')[2]
        except:
            return data
        if task_type == ['其他']:
            now = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
            task_status = ele.xpath("//*[@id='legendBasic']/table/tbody/tr[6]/td/span/text()")
            if task_status == [' 已完成'] and pro_time == now:
                data['status'] = '通过'
            if task_status == [' 进行中'] and pro_time == now:
                data['status'] = '未通过'
        return data


    def zentao_get(self, api_path):
        response = requests.get(self.host + api_path, self.params)
        # print(response.content)
        return response.text

    # def zentao_post(self, api_path, data=None, json_data=None):
    #     response = requests.post(self.host + api_path, data=data, json=json_data, params=self.params)
    #
    #     return self._get_zentao_result(response.json())


class Online_Xlsl(Zentao):

    def __init__(self, host, account, password):
        super(Online_Xlsl, self).__init__(host=host, account=account, password=password)
        self.project_data = self.get_task_list()

    def create_online_xlsl(self):
        """
        制作文档
        :return:
        """
        pre_insert_datas, pro_insert_datas = self.get_online_data()
        # 上线结果表
        pro_time = datetime.datetime.now().strftime("%Y%m%d")
        pro_name = pro_time + '上线结果报告'
        pro_wb = Workbook()
        pro_ws = pro_wb.create_sheet(index=0, title=pro_name)
        # 设置表头
        pro_ws.merge_cells('A1:J1')
        pro_ws.cell(1, 1).value = pro_name
        # 设置第一行
        pro_ws.append(['上线内容', '上线版本', '上线结果'])
        for pro_insert_data in pro_insert_datas:
            pro_ws.append(pro_insert_data)
        pro_wb.save(pro_name + '.xlsx')

        # 预上线文档
        # 预上线是下一天上线
        pre_time = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%Y%m%d")
        pre_name = pre_time + '上线内容测试报告'
        pre_wb = Workbook()
        pre_ws = pre_wb.create_sheet(index=0, title=pre_name)
        # 设置表头
        pre_ws.merge_cells('A1:J1')
        pre_ws.cell(1, 1).value = pre_name
        # # 设置表头样式
        # font = Font(name=u'宋体', bold=True)
        # align = Alignment(horizontal='center', vertical='center')
        # ws.font = font
        # ws.align = align
        # 设置第一行
        pre_ws.append(['产品线', '测试内容', '系统', '测试结果', '测试人', '研发', '配置审核', '版本', '复测结果', '复测人'])
        for pre_insert_data in pre_insert_datas:
            pre_ws.append(pre_insert_data)
        pre_wb.save(pre_name + '.xlsx')
        return True

    def get_online_data(self):
        """
        获取文档内容。
        :param project_data:
        :return:
        """
        pre_insert_datas = []
        pro_insert_datas = []
        for project_id, task_datas in self.project_data.items():
            for task_id, task_data in task_datas.items():
                # 对禅道获取到的数据进行判断
                if str(task_data.get('status')) == ' 进行中':
                    pre_online_remaek, online_remaek = self.get_subtask_ids_by_task_id(task_id)
                    # 任务已经是完成状态，获取有准备上线备注的
                    if pre_online_remaek:
                        test_content = 'taskid=' + str(task_id) + '  ' + task_data.get('task_name')
                        # 预上线文档生成
                        # 任务有上线备注，若最后修改时间为今天或者上线备注为空(上线失败)，则该任务预计明天上线
                        for pre_task_id, pre_task_data in pre_online_remaek.items():
                            online_project_name = pre_task_data.get('project', '')
                            version = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%Y%m%d")
                            insert_data = [task_data.get('product_line'), test_content, online_project_name, '通过',
                                       pre_task_data.get('test_people'),
                                       pre_task_data.get('develop_people'), pre_task_data.get('configure'),
                                       version, '', pre_task_data.get('test_people')]
                            print(insert_data, 'pre')
                            pre_insert_datas.append(insert_data)
                        # 上线结果文档
                        # 首先获取前一天需要上线的task
                            # 如果有上线备注，则上线成功，否则失败
                if str(task_data.get('status')) == ' 已完成':
                    pre_online_remaek, online_remaek = self.get_subtask_ids_by_task_id(task_id)
                    if online_remaek:
                        test_content = 'taskid=' + str(task_id) + '  ' + task_data.get('task_name')
                        # 预上线文档生成
                        # 任务有上线备注，若最后修改时间为今天或者上线备注为空(上线失败)，则该任务预计明天上线
                        for pro_task_id, pro_task_data in online_remaek.items():
                            online_project_name = pro_task_data.get('project', '')
                            version = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%Y%m%d")
                            insert_data = [task_data.get('product_line'), test_content, online_project_name, '通过',
                                       pro_task_data.get('test_people'),
                                       pro_task_data.get('develop_people'), pro_task_data.get('configure'),
                                       version, pro_task_data.get('result'), pro_task_data.get('test_people')]
                            print(insert_data, 'pro')
                            pro_insert_datas.append(insert_data)
        return pre_insert_datas, pro_insert_datas[:1]


class Messenger(object):
    MAIL_SERVICE_URL = r'http://10.13.225.228:8082/api/mail/'
    def post_email(self, to, subjects, content, type="6"):
        """
        发送邮件
        """
        print('开始发送邮件')
        try:
            email_body = {"inboxes": to, "subject": subjects, "content": content, "module": "mn_alarm", "type": 6}
            r = requests.post(self.MAIL_SERVICE_URL, data=json.dumps(email_body))
            r.raise_for_status()
            result = r.json()
            if result['code'] == 0:
                print('send mail to %s OK' % to)
                return 0
            else:
                print('send mail to %s Failed, content =%s, Msg:%s' % (to, content, result))
                return -1
        except Exception as e:
            print('On send mail err occurred:%s' % e)
            return -1


def post_pro_email():
    # 发上线文档
    pro_time = datetime.datetime.now().strftime('%Y%m%d')
    pro_name = '系统更新-上线结果通知[' + str(pro_time) + ']'
    receiver = 'zening.jin@capitalonline.net'
    # receiver = 'tao.xu@capitalonline.net'

    content = """
    <br><br>
     各位同事：
    <br><br><br>
    大家好，今日上线结果通知，请支持，谢谢~
    <br>
    """
    content += '统计时间:' + pro_time
    content += """
        <br>
        <table border="1" cellspacing="0" width="80%" height="150" style="margin-top: 30px" align="left">
        <caption style="font-size: 20px">上线内容测试报告</caption>
        <tr height='3%',class='tabletitle'>
        <th width='20%'>产品线</th>
        <th width="20%">测试内容</th>
        <th width="20%">系统</th>
        <th width="5%">测试结果</th>
        <th width="5%">测试人</th>
        <th width="5%">研发</th>
        <th width="5%">配置审核</th>
        <th width="5%">版本</th>
        <th width="5%">复测结果</th>
        <th width="5%">复测人</th>
        </tr>
    """
    for pro_insert_data in pro_insert_datas:
        content += """
            <tr>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            </tr>
        """.format(pro_insert_data[0], pro_insert_data[1], pro_insert_data[2],
                   pro_insert_data[3], pro_insert_data[4], pro_insert_data[5],
                   pro_insert_data[6], pro_insert_data[7], pro_insert_data[8],
                   pro_insert_data[9])


    subject = pro_name
    a = Messenger()
    a.post_email(receiver, subject, content)


def post_pre_email():
    # 发上线文档
    pre_time = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y%m%d')
    content_time = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%Y年%m月%d日')
    pre_name = '系统更新--业务平台日常更新[' + str(pre_time) + ']'
    receiver = 'zening.jin@capitalonline.net'
    # receiver = 'tao.xu@capitalonline.net'
    content = """
        <br><br>
        各位同事：
        <br><br>
        大家好，GIC更新，请支持，谢谢~
        <br><br><br>
        """
    content += '更新时间' + content_time
    content += """
        更新方式：日常更新
        <br>
        运维人员：郝志新【18910538051】
        <br><br>
        请各产品线确认已通过上线评审。
        <br>
        请通过测试的产品线统一准备好上线文档，提交应用运维。（10月12日10点准时开始测试）
        <br>
        若产品线有需要追加的任务请负责人回复邮件，谢谢~
        <br>
        如有其他异议或疑问请及时回复本邮件。
        <br>
        """
    content += '以下是' + pre_time + '上线内容'
    content += """
        <br>
        <table border="1" cellspacing="0" width="80%" height="150" style="margin-top: 30px" align="left">
        <caption style="font-size: 20px">上线内容测试报告</caption>
        <tr height='3%',class='tabletitle'>
        <th width='20%'>产品线</th>
        <th width="20%">测试内容</th>
        <th width="20%">系统</th>
        <th width="5%">测试结果</th>
        <th width="5%">测试人</th>
        <th width="5%">研发</th>
        <th width="5%">配置审核</th>
        <th width="5%">版本</th>
        <th width="5%">复测结果</th>
        <th width="5%">复测人</th>
        </tr>
    """
    for pre_insert_data in pre_insert_datas:
        content += """
            <tr>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            <td align='center'>{}</td>
            </tr>
        """.format(pre_insert_data[0], pre_insert_data[1], pre_insert_data[2],
                   pre_insert_data[3], pre_insert_data[4], pre_insert_data[5],
                   pre_insert_data[6], pre_insert_data[7], pre_insert_data[8],
                   pre_insert_data[9])

    subject = pre_name
    a = Messenger()
    a.post_email(receiver, subject, content)


# 爬虫获取数据
logging.info('我来啦')
zentao_host = "http://cds-cdproject.capitalonline.net"
zentao_account = "jinzn"
zentao_password = "zaq123edc"
test_object = Online_Xlsl(host=zentao_host, account=zentao_account, password=zentao_password)
# 生成表格查看
test_object.create_online_xlsl()
# 发送邮件
pre_insert_datas, pro_insert_datas = test_object.get_online_data()
post_pro_email()
post_pre_email()

# # 30 8 * * *
# crontab -e
# service cron restart
